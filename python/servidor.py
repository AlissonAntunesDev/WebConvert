from os import  listdir,path
from re import search,findall,DOTALL,MULTILINE,IGNORECASE
from PyPDF2 import PdfReader
from pandas import DataFrame
from openpyxl.styles import Alignment
from openpyxl import load_workbook

# Função para extrair o número SEI do texto
def extrair_numero_sei(string):
    padrao = r'\d+\.\d+/\d+-\d+'
    resultado = search(padrao, string)
    return resultado.group() if resultado else "não encontrado"

# Função para remover quebras de linha em todas as entradas de um dicionário
def remover_quebras_de_linha_em_dicionario(dicionario):
    for chave, valores in dicionario.items():
        dicionario[chave] = [valor.replace('\n', '') for valor in valores]
    return dicionario

# Função para extrair texto de um PDF
def extrair_texto_pdf(caminho_pdf):
    with open(caminho_pdf, "rb") as arquivo:
        leitor = PdfReader(arquivo)
        print("".join(pagina.extract_text() + "\n" for pagina in leitor.pages))    
        return "".join(pagina.extract_text() + "\n" for pagina in leitor.pages)

# Função para extrair nome do parecer, processo SEI e ementa
def extrair_parecer_ementa(texto):
    linhas = texto.strip().split("\n")
    nome_parecer = next((linha.strip() for linha in linhas if "PARECER" in linha), "")
    ementa_linhas = []
    for linha in linhas[linhas.index(nome_parecer) + 1:]:
        ementa_linhas.append(linha.strip())
        if linha.startswith("Processo SEI"):
            processo_sei = extrair_numero_sei(linha)
            break
    ementa = " ".join(ementa_linhas).strip()
    processo_sei = extrair_numero_sei(ementa)
    return nome_parecer, processo_sei, ementa

# # Função para reduzir a ementa, removendo o texto antes da primeira quebra de linha 
# def reduzir_ementa(string):
#     # Encontra a posição da primeira quebra de linha
#     pos_quebra = string.find('\n')
    
#     # Se houver uma quebra de linha, retorna o texto após ela
#     if pos_quebra != -1:
#         return string[pos_quebra + 1:].strip()
#     else:
#         return string  # Caso não haja quebra de linha
    

# Função para extrair assinaturas completas
def extrair_assinaturas_completas(texto):
    # Regex para capturar nome, cargo e data, considerando quebras de linha
    padrao = r"Documento assinado eletronicamente por\s+([\wÀ-ÿ\s]+?),\s*\n*([\wÀ-ÿ\s\(\)-]+?),\s*\n*em\s*(\d{2}/\d{2}/\d{4})"
    matches = findall(padrao, texto, DOTALL | MULTILINE)
    
    assinaturas_limpas = []
    for nome, cargo, data in matches:
        # Remove quebras de linha e espaços extras no nome e no cargo
        nome_limpo = " ".join(nome.split())
        cargo_limpo = " ".join(cargo.split())
        # print(f"=== {nome_limpo}, {cargo_limpo}, {data}")
        assinaturas_limpas.append((nome_limpo, cargo_limpo, data))
    
    return assinaturas_limpas

# Função para buscar a classificação LAI no texto
def buscar_classificacao_lai(texto):
    padrao = r'\b(ATO PREPARATÓRIO|SIGILO PROFISSIONAL|DOCUMENTO PÚBLICO)\b'
    resultado = search(padrao, texto, IGNORECASE)
    return resultado.group(0) if resultado else "Não encontrada"

# Função para ajustar a planilha Excel
def ajustar_planilha(caminho_arquivo):
    print("Ajustando o arquivo Excel ...")

    wb = load_workbook(caminho_arquivo)
    sheet = wb.active

    # Ajusta a largura de todas as colunas com base no conteúdo
    for col in sheet.columns:
        max_length = 0
        col_letter = col[0].column_letter  # Pega a letra da coluna
        
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
                # Centraliza e ativa a quebra de linha em todas as células
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # Define a largura com limite na coluna N
        largura = max_length + 2
        if col_letter == 'N':
            largura = min(largura, 200)

        
        sheet.column_dimensions[col_letter].width = largura
        sheet.column_dimensions['O'].width = 30  # Ajusta a largura da coluna O
        

    # Ajusta a altura da primeira linha
    sheet.row_dimensions[1].height = 40

    # Ajusta a altura das células da coluna N automaticamente
    for row in sheet.iter_rows(min_row=2):
        n_cell = row[13]  # Coluna N (índice 13, começando em 0)
        if n_cell.value:
            # Calcula a altura com base no comprimento do texto
            linhas = len(str(n_cell.value)) // 30 + 1  # Aproximadamente 30 caracteres por linha
            sheet.row_dimensions[n_cell.row].height = linhas * 3

    # Salva as alterações
    wb.save(caminho_arquivo)
    
    print("Ajustes concluídos com sucesso!")

# Inicializa o dicionário para armazenar os pareceres
pareceres = {campo: [] for campo in [
    "NÚMERO E ANO DE REGISTRO DO PARECER (COM LINK)", "PROCESSO SEI", "DATA DO PARECER",
    "DATA APROVAÇÃO COORDENADOR", "DATA APROVAÇÃO ADJUNTO", "DATA APROVAÇÃO PROCURADOR-GERAL",
    "DATA APROVAÇÃO MINISTRO (SIM ou NÃO)", "DATA APROVAÇÃO PRESIDENTE DA REPÚBLICA (SIM ou NÃO)",
    "PROCURADOR AUTOR DO PARECER", "COORDENADOR APROVOU PARECER", "ADJUNTO APROVOU PARECER",
    "APROVAÇÃO PROCURADOR-GERAL (SIM OU NÃO)", "APROVAÇÃO MINISTRO (SIM OU NÃO)", "EMENTA",
    "CONSULENTE", "CLASSIFICAÇÃO LAI"]
}

# Processa cada arquivo PDF na pasta atual
caminho_pasta = "./"
pareceres = {
    "NÚMERO E ANO DE REGISTRO DO PARECER (COM LINK)": [],
    "PROCESSO SEI": [],
    "DATA DO PARECER": [],
    "DATA APROVAÇÃO COORDENADOR": [],
    "DATA APROVAÇÃO ADJUNTO": [],	
    "DATA APROVAÇÃO PROCURADOR-GERAL": [],	
    "DATA APROVAÇÃO MINISTRO (SIM ou NÃO)": [],
    "DATA APROVAÇÃO PRESIDENTE DA REPÚBLICA (SIM ou NÃO)": [],	
    "PROCURADOR AUTOR DO PARECER": [],	
    "COORDENADOR APROVOU PARECER": [],	
    "ADJUNTO APROVOU PARECER": [],	
    "APROVAÇÃO PROCURADOR-GERAL (SIM OU NÃO)": [],
    "APROVAÇÃO MINISTRO (SIM OU NÃO)": [],
    "EMENTA": [],
    "CONSULENTE": [],
    "CLASSIFICAÇÃO LAI": [],
}

for arquivo in listdir(caminho_pasta):
    if arquivo.endswith(".pdf"):
        caminho_pdf = path.join(caminho_pasta, arquivo)
        print(f"Processando {caminho_pdf} ...")
        print(f"====================================================================================================================================================================")
        texto_pdf = extrair_texto_pdf(caminho_pdf)        
        
        nome, processo_sei, ementa = extrair_parecer_ementa(texto_pdf)
        assinaturas_extraidas = extrair_assinaturas_completas(texto_pdf)
        
        procurador = coordenador = adjunto = procurador_geral = ministro = presidente = "NÃO"
        data_procurador = data_coordenador = data_adjunto = data_procurador_geral = data_ministro = data_presidente = "NÃO"
        classificacao_lai = buscar_classificacao_lai(ementa)

        for nome_ass, cargo, data in assinaturas_extraidas:
            if "Procurador(a)-Geral" in cargo:
                if ("Procurador(a)-Geral Adjunto(a)" or "Procurador(a)-Geral Adjunto(a) Substituto(a)")in cargo:
                    adjunto, data_adjunto = nome_ass, data
                else:
                    procurador_geral, data_procurador_geral = nome_ass, data
            elif ("Coordenador(a)" or "Coordenador(a) Substituto(a)") in cargo:
                coordenador, data_coordenador = nome_ass, data
            elif ("Procurador(a) da Fazenda Nacional" or "Procurador(a) da Fazenda Nacional Substituto(a)") in cargo:
                if procurador == "NÃO":
                    procurador, data_procurador = nome_ass, data
            elif ("Ministro(a)" or "Ministro(a) Substituto(a)") in cargo:
                ministro, data_ministro = nome_ass, data
            elif ("Presidente da República" or "Presidente da República Substituto(a)") in cargo:
                presidente, data_presidente = nome_ass, data
                
        pareceres["NÚMERO E ANO DE REGISTRO DO PARECER (COM LINK)"].append(nome)
        pareceres["PROCESSO SEI"].append(processo_sei)
        pareceres["DATA DO PARECER"].append(data_procurador)
        pareceres["DATA APROVAÇÃO COORDENADOR"].append(data_coordenador)
        pareceres["DATA APROVAÇÃO ADJUNTO"].append(data_adjunto)
        pareceres["DATA APROVAÇÃO PROCURADOR-GERAL"].append(data_procurador_geral)
        pareceres["DATA APROVAÇÃO MINISTRO (SIM ou NÃO)"].append(data_ministro)
        pareceres["DATA APROVAÇÃO PRESIDENTE DA REPÚBLICA (SIM ou NÃO)"].append(data_presidente)
        pareceres["PROCURADOR AUTOR DO PARECER"].append(procurador)
        pareceres["COORDENADOR APROVOU PARECER"].append(coordenador)
        pareceres["ADJUNTO APROVOU PARECER"].append(adjunto)
        pareceres["APROVAÇÃO PROCURADOR-GERAL (SIM OU NÃO)"].append(procurador_geral)
        pareceres["APROVAÇÃO MINISTRO (SIM OU NÃO)"].append(ministro)
        pareceres["EMENTA"].append(ementa)
        pareceres["CONSULENTE"].append("         ")
        pareceres["CLASSIFICAÇÃO LAI"].append(classificacao_lai)

pareceres = remover_quebras_de_linha_em_dicionario(pareceres)

# Encontrar o tamanho máximo das listas
tamanho_maximo = max(len(valores) for valores in pareceres.values())

# # Salva os dados em um arquivo Excel
# df = DataFrame(pareceres)
# caminho_arquivo = "Pareceres.xlsx"
# df.to_excel(caminho_arquivo, index=False)

# # Ajusta layout do arquivo Excel
# ajustar_planilha(caminho_arquivo)

